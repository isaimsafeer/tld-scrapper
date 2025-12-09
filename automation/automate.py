from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from typing import Dict, Any, List
import logging
from datetime import datetime
import time
import openpyxl
from openpyxl import Workbook
import re
import json

logger = logging.getLogger(__name__)


def setup_driver(headless: bool = False) -> webdriver.Chrome:
    """
    Setup Chrome WebDriver with options
    
    Args:
        headless: Run browser in headless mode
        
    Returns:
        Configured Chrome WebDriver instance
    """
    chrome_options = Options()
    
    logger.info(f"Setting up driver with headless={headless}")
    
    if headless:
        chrome_options.add_argument("--headless")
        logger.info("Running in HEADLESS mode")
    else:
        logger.info("Running in VISIBLE mode (browser should open)")
    
    # Additional options for stability
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--window-size=1920,1080")
    chrome_options.add_argument("--start-maximized")
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option("useAutomationExtension", False)
    
    # Ensure no headless flags are accidentally set
    chrome_options.add_experimental_option("detach", True)
    
    # Initialize driver
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=chrome_options)
    
    return driver


def extract_registrar_from_url(url: str) -> str:
    """
    Extract registrar name from URL
    
    Args:
        url: Full URL of the page
        
    Returns:
        Registrar name (domain name without TLD)
    """
    # Extract domain from URL
    match = re.search(r'https?://(?:www\.)?([^/]+)', url)
    if match:
        domain = match.group(1)
        # Get the main domain name (e.g., 'porkbun' from 'porkbun.com')
        parts = domain.split('.')
        if len(parts) >= 2:
            return parts[0]
        return domain
    return "unknown"


def parse_price_data(cell_value: str) -> Dict[str, Any]:
    """
    Parse price information from cell value
    
    Args:
        cell_value: Raw cell value containing registrar, promo, and price
        
    Returns:
        Dictionary with parsed price information
    """
    result = {
        "registrar": "",
        "price": None,
        "currency": "USD",
        "pricing_notes": "",
        "promo_code": ""
    }
    
    if not cell_value or cell_value.strip() == "":
        return result
    
    # Debug: Log the raw cell value
    logger.debug(f"Parsing cell value: {repr(cell_value[:100])}")
    
    lines = cell_value.strip().split('\n')
    
    # First line is usually the registrar
    if lines:
        result["registrar"] = lines[0].strip().lower()
        logger.debug(f"Extracted registrar: {result['registrar']}")
    
    # Look for promo code
    promo_match = re.search(r'Promo code\s+(\S+)', cell_value, re.IGNORECASE)
    if promo_match:
        result["promo_code"] = promo_match.group(1)
        result["pricing_notes"] = f"promo code: {promo_match.group(1)}"
        logger.debug(f"Found promo code: {result['promo_code']}")
    
    # Look for price with currency symbol
    # Handles formats like: $9.99, £9.99, €9.99, £➜ $12.71
    price_patterns = [
        (r'\$(\d+\.?\d*)', 'USD'),  # Dollar
        (r'£➜\s*\$(\d+\.?\d*)', 'USD'),  # Pound converted to dollar  
        (r'£(\d+\.?\d*)', 'GBP'),  # Pound
        (r'€(\d+\.?\d*)', 'EUR'),  # Euro
    ]
    
    for pattern, currency in price_patterns:
        price_match = re.search(pattern, cell_value)
        if price_match:
            result["price"] = float(price_match.group(1))
            result["currency"] = currency
            logger.debug(f"Found price: {result['price']} {currency} using pattern {pattern}")
            break
    
    if result["price"] is None:
        logger.warning(f"Could not extract price from: {repr(cell_value[:100])}")
    
    # Check for special notes like "registration" or "renewal"
    if "registration" in cell_value.lower():
        if result["pricing_notes"]:
            result["pricing_notes"] += ", registration price"
        else:
            result["pricing_notes"] = "registration price"
    
    if "renewal" in cell_value.lower() and "registration" in cell_value.lower():
        # This is a multi-year value with different prices
        if result["pricing_notes"]:
            result["pricing_notes"] += ", includes renewal"
        else:
            result["pricing_notes"] = "includes renewal"
    
    return result


def map_column_to_operation(column_name: str) -> str:
    """
    Map column header to operation type
    
    Args:
        column_name: Column header text
        
    Returns:
        Operation type (register, renew, transfer)
    """
    column_lower = column_name.lower()
    
    if "registration" in column_lower or "register" in column_lower:
        return "register"
    elif "renewal" in column_lower or "renew" in column_lower:
        return "renew"
    elif "transfer" in column_lower:
        return "transfer"
    elif "value" in column_lower or "year" in column_lower:
        # Special case for "Best 3 Year Value"
        return "multi_year"
    else:
        return "unknown"


def process_excel_to_structured_data(excel_filename: str, url: str, task_id: str) -> List[Dict[str, Any]]:
    """
    Process Excel file and convert to structured format
    
    Args:
        excel_filename: Path to Excel file
        url: Source URL for registrar extraction
        task_id: Task identifier
        
    Returns:
        List of structured data dictionaries
    """
    structured_data = []
    scraped_at = datetime.now().isoformat()
    registrar_from_url = extract_registrar_from_url(url)
    
    try:
        # Load the Excel file
        workbook = openpyxl.load_workbook(excel_filename)
        sheet = workbook.active
        
        # Get headers (first row) - include ALL cells
        headers = []
        for cell in sheet[1]:
            headers.append(str(cell.value).strip() if cell.value else "")
        
        logger.info(f"Task {task_id}: Processing {len(headers)} columns")
        logger.info(f"Task {task_id}: Headers: {headers}")
        
        # Process each data row (skip header row)
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not any(row):  # Skip empty rows
                continue
            
            # According to your example, column 0 is empty and column 1 has the TLD
            # Let's check both first and second columns for TLD
            tld = ""
            tld_col_idx = -1
            
            # Check first column
            if row[0] and str(row[0]).strip():
                tld = str(row[0]).strip()
                tld_col_idx = 0
            # If first column is empty, check second column
            elif len(row) > 1 and row[1] and str(row[1]).strip():
                tld = str(row[1]).strip()
                tld_col_idx = 1
            
            if not tld or not tld.startswith('.'):
                logger.warning(f"Task {task_id}: Row {row_idx} has no valid TLD (found: '{row[0]}', '{row[1] if len(row) > 1 else 'N/A'}'), skipping")
                continue
            
            logger.info(f"Task {task_id}: Processing row {row_idx} - TLD: {tld}")
            
            # Process each column after the TLD column
            start_col = tld_col_idx + 1
            for col_idx in range(start_col, len(row)):
                if col_idx >= len(headers):
                    break
                
                cell_value = row[col_idx]
                column_name = headers[col_idx]
                
                # Skip if column name or cell value is empty
                if not column_name or not cell_value or str(cell_value).strip() == "":
                    continue
                
                operation = map_column_to_operation(column_name)
                
                # Parse price data from cell
                price_info = parse_price_data(str(cell_value))
                
                if price_info["price"] is None:
                    logger.warning(f"Task {task_id}: No price found in row {row_idx}, column '{column_name}' - Value: '{str(cell_value)[:50]}'")
                    continue
                
                # Create structured record
                record = {
                    "tld": tld,
                    "registrar": price_info["registrar"] if price_info["registrar"] else registrar_from_url,
                    "operation": operation,
                    "price": price_info["price"],
                    "currency": price_info["currency"],
                    "years": 3 if operation == "multi_year" else 1,
                    "pricing_notes": price_info["pricing_notes"],
                    "promo_code": price_info["promo_code"],
                    "scraped_at": scraped_at,
                    "source_url": url,
                    "column_name": column_name
                }
                
                structured_data.append(record)
                logger.info(f"Task {task_id}: Added record - TLD: {tld}, Registrar: {record['registrar']}, Operation: {operation}, Price: {price_info['price']}")
        
        logger.info(f"Task {task_id}: Successfully processed {len(structured_data)} records from Excel")
        
    except Exception as e:
        logger.error(f"Task {task_id}: Error processing Excel file: {e}")
        import traceback
        logger.error(f"Task {task_id}: Traceback: {traceback.format_exc()}")
    
    return structured_data


def save_structured_data(structured_data: List[Dict[str, Any]], task_id: str) -> str:
    """
    Save structured data to JSON and Excel files
    
    Args:
        structured_data: List of structured records
        task_id: Task identifier
        
    Returns:
        Path to JSON file
    """
    # Save as JSON
    json_filename = f"structured_data_{task_id}.json"
    with open(json_filename, 'w', encoding='utf-8') as f:
        json.dump(structured_data, f, indent=2, ensure_ascii=False)
    logger.info(f"Task {task_id}: Structured data saved to {json_filename}")
    
    # Save as Excel
    excel_filename = f"structured_data_{task_id}.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Structured Data"
    
    # Write headers
    if structured_data:
        headers = list(structured_data[0].keys())
        sheet.append(headers)
        
        # Write data
        for record in structured_data:
            row = [record.get(header, "") for header in headers]
            sheet.append(row)
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    workbook.save(excel_filename)
    logger.info(f"Task {task_id}: Structured data saved to {excel_filename}")
    
    return json_filename


def run_automation(task_id: str, url: str, timeout: int, headless: bool) -> Dict[str, Any]:
    """
    Main automation function
    
    Args:
        task_id: Unique task identifier
        url: Target website URL
        timeout: Maximum wait time for elements
        headless: Run in headless mode
        
    Returns:
        Dictionary containing automation results or error information
    """
    driver = None
    
    try:
        logger.info(f"Task {task_id}: Starting automation for {url}")
        
        # Setup driver
        driver = setup_driver(headless=headless)
        
        # Navigate to URL
        driver.get(url)
        logger.info(f"Task {task_id}: Navigated to {url}")
        
        # Wait for page to load
        wait = WebDriverWait(driver, timeout)
        
        # Wait for body element to be present
        wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))
        
        # ===== EXTRACT TABLE DATA AND SAVE TO EXCEL =====
        
        page_title = driver.title
        logger.info(f"Task {task_id}: Page title: {page_title}")
        
        # Wait for the specific table to load
        logger.info(f"Task {task_id}: Waiting for table to load...")
        table = wait.until(EC.presence_of_element_located(
            (By.XPATH, "(//table[@id='DataTables_Table_0'])[1]")
        ))
        logger.info(f"Task {task_id}: Table found!")
        
        # Extract table headers from the correct location
        headers = []
        try:
            # Get headers from the dataTables_scrollHeadInner div
            header_cells = driver.find_elements(By.XPATH, "//div[contains(@class,'dataTables_scrollHeadInner')]//thead//tr//th")
            headers = [cell.text.strip() for cell in header_cells if cell.text.strip()]
            
            # Add empty first column to align with data rows that have domain in first column
            if headers:
                headers.insert(0, "Domain")  # Insert "Domain" as first column header
                logger.info(f"Task {task_id}: Found {len(headers)} headers (with Domain added): {headers}")
            else:
                logger.info(f"Task {task_id}: Found {len(headers)} headers: {headers}")
        except Exception as e:
            logger.warning(f"Task {task_id}: Could not extract headers: {e}")
            headers = []
        
        # Extract table rows from multiple pages (pagination)
        rows_data = []
        
        # Find the second-last <li> which contains the last page number
        last_page_element = driver.find_element(By.XPATH, "//div[@id='DataTables_Table_0_paginate']//li[last()-1]")

        # Extract the text and convert to integer
        pages_to_scrape = int(last_page_element.text)

        print("Total pages:", pages_to_scrape)
        
        try:
            for page_num in range(pages_to_scrape):
                logger.info(f"Task {task_id}: Extracting data from page {page_num + 1}...")
                
                # Wait a bit for the page to load
                time.sleep(2)
                
                # Get all rows from tbody of the first table
                table_rows = driver.find_elements(By.XPATH, "(//table[@id='DataTables_Table_0'])[1]//tbody//tr")
                logger.info(f"Task {task_id}: Found {len(table_rows)} rows on page {page_num + 1}")
                
                for idx, row in enumerate(table_rows):
                    try:
                        # Get the row ID (domain name)
                        row_id = row.get_attribute('id')
                        
                        # Extract all cells in this row
                        cells = row.find_elements(By.TAG_NAME, "td")
                        row_data = [cell.text.strip() for cell in cells]
                        
                        if row_data:  # Only add non-empty rows
                            rows_data.append(row_data)
                            logger.debug(f"Task {task_id}: Row {idx} (id={row_id}): {row_data[:3]}...")  # Log first 3 cells
                        
                    except Exception as row_error:
                        logger.warning(f"Task {task_id}: Error extracting row {idx}: {row_error}")
                        continue
                
                logger.info(f"Task {task_id}: Extracted {len(table_rows)} rows from page {page_num + 1}. Total rows so far: {len(rows_data)}")
                
                # Click Next button if not on the last page
                if page_num < pages_to_scrape - 1:
                    try:
                        next_button = driver.find_element(By.XPATH, "//*[@id='DataTables_Table_0_next']/a")
                        
                        # Check if the Next button is disabled
                        parent_li = driver.find_element(By.XPATH, "//*[@id='DataTables_Table_0_next']")
                        if "disabled" in parent_li.get_attribute("class"):
                            logger.info(f"Task {task_id}: Next button is disabled. No more pages available.")
                            break
                        
                        # Scroll to button and click
                        driver.execute_script("arguments[0].scrollIntoView(true);", next_button)
                        time.sleep(1)
                        next_button.click()
                        logger.info(f"Task {task_id}: Clicked Next button to go to page {page_num + 2}")
                        
                        # Wait for new page to load
                        time.sleep(2)
                        
                    except Exception as click_error:
                        logger.warning(f"Task {task_id}: Could not click Next button: {click_error}")
                        break
            
            logger.info(f"Task {task_id}: Successfully extracted {len(rows_data)} total data rows from {page_num + 1} pages")
        except Exception as e:
            logger.error(f"Task {task_id}: Error extracting rows: {e}")
        
        # Create Excel file with raw data
        excel_filename = f"table_data_{task_id}.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Table Data"
        
        # Write headers
        if headers:
            sheet.append(headers)
            logger.info(f"Task {task_id}: Headers written to Excel")
        
        # Write data rows
        for row_data in rows_data:
            sheet.append(row_data)
        
        # Auto-adjust column widths for better readability
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 for very long content
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save Excel file
        workbook.save(excel_filename)
        logger.info(f"Task {task_id}: Raw data saved to {excel_filename}")
        
        # ===== PROCESS DATA INTO STRUCTURED FORMAT =====
        logger.info(f"Task {task_id}: Processing data into structured format...")
        structured_data = process_excel_to_structured_data(excel_filename, url, task_id)
        
        # Save structured data
        json_filename = save_structured_data(structured_data, task_id)
        
        # Keep browser open for 3 seconds so you can see it
        logger.info(f"Task {task_id}: Browser will remain open for 3 seconds...")
        time.sleep(3)
        
        # ===== END OF AUTOMATION LOGIC =====
        
        # Prepare results
        result = {
            "status": "completed",
            "result": {
                "page_title": page_title,
                "url": driver.current_url,
                "raw_excel_file": excel_filename,
                "structured_json_file": json_filename,
                "structured_excel_file": f"structured_data_{task_id}.xlsx",
                "rows_extracted": len(rows_data),
                "structured_records": len(structured_data),
                "columns": len(headers) if headers else 0,
                "headers": headers,
                "completed_at": datetime.now().isoformat()
            }
        }
        
        logger.info(f"Task {task_id}: Automation completed successfully")
        logger.info(f"Task {task_id}: Created {len(structured_data)} structured records")
        return result
        
    except Exception as e:
        logger.error(f"Task {task_id}: Error - {str(e)}")
        import traceback
        logger.error(f"Task {task_id}: Traceback - {traceback.format_exc()}")
        return {
            "status": "failed",
            "error": str(e),
            "traceback": traceback.format_exc()
        }
        
    finally:
        if driver:
            driver.quit()
            logger.info(f"Task {task_id}: Browser closed")